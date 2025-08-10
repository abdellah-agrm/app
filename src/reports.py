# reports.py
import tkinter as tk
from tkinter import ttk, messagebox
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkcalendar import DateEntry
from database import create_connection
import sqlite3
from datetime import datetime
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment

class ReportsManager:
    def __init__(self, parent, user_data):
        self.parent = parent
        self.user_data = user_data
        
        # Create main frames
        self.main_frame = ttk.Frame(self.parent)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Notebook for different report types
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # Sales Reports Tab
        self.sales_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.sales_tab, text="Sales Reports")
        
        # Stock Reports Tab
        self.stock_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.stock_tab, text="Stock Reports")
        
        # Initialize UI
        self.create_sales_report_ui()
        self.create_stock_report_ui()
        
    def create_sales_report_ui(self):
        # Date range frame
        date_frame = ttk.Frame(self.sales_tab)
        date_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(date_frame, text="From:").pack(side=tk.LEFT, padx=5)
        self.sales_date_from = DateEntry(date_frame, width=12)
        self.sales_date_from.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(date_frame, text="To:").pack(side=tk.LEFT, padx=5)
        self.sales_date_to = DateEntry(date_frame, width=12)
        self.sales_date_to.pack(side=tk.LEFT, padx=5)
        
        # Set default dates (last 30 days)
        self.sales_date_from.set_date(datetime.now().replace(day=1))
        self.sales_date_to.set_date(datetime.now())
        
        # Group by frame
        group_frame = ttk.Frame(self.sales_tab)
        group_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(group_frame, text="Group By:").pack(side=tk.LEFT, padx=5)
        self.sales_group_by = ttk.Combobox(group_frame, values=['Day', 'Week', 'Month', 'Year', 'Product', 'Payment Method'])
        self.sales_group_by.pack(side=tk.LEFT, padx=5)
        self.sales_group_by.current(0)
        
        # Buttons frame
        buttons_frame = ttk.Frame(self.sales_tab)
        buttons_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(buttons_frame, text="Generate Report", command=self.generate_sales_report,
                  style='primary.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(buttons_frame, text="Export to PDF", command=self.export_sales_pdf,
                  style='danger.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(buttons_frame, text="Export to Excel", command=self.export_sales_excel,
                  style='success.TButton').pack(side=tk.LEFT, padx=5)
        
        # Results frame
        results_frame = ttk.Frame(self.sales_tab)
        results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Treeview
        self.sales_tree = ttk.Treeview(results_frame)
        self.sales_tree.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=self.sales_tree.yview)
        self.sales_tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Summary frame
        summary_frame = ttk.Frame(self.sales_tab)
        summary_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(summary_frame, text="Total Sales:").pack(side=tk.LEFT, padx=5)
        self.total_sales_var = tk.StringVar(value="$0.00")
        ttk.Label(summary_frame, textvariable=self.total_sales_var, font=('Helvetica', 12, 'bold')).pack(side=tk.LEFT, padx=5)
        
    def create_stock_report_ui(self):
        # Filter frame
        filter_frame = ttk.Frame(self.stock_tab)
        filter_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(filter_frame, text="Filter:").pack(side=tk.LEFT, padx=5)
        self.stock_filter = ttk.Combobox(filter_frame, values=['All', 'Low Stock (<5)', 'By Brand'])
        self.stock_filter.pack(side=tk.LEFT, padx=5)
        self.stock_filter.current(0)
        
        ttk.Label(filter_frame, text="Brand:").pack(side=tk.LEFT, padx=5)
        self.stock_brand = ttk.Combobox(filter_frame)
        self.stock_brand.pack(side=tk.LEFT, padx=5)
        self.stock_brand['state'] = 'disabled'
        
        # Load brands
        self.load_brands()
        
        # Bind filter change
        self.stock_filter.bind('<<ComboboxSelected>>', self.on_stock_filter_change)
        
        # Buttons frame
        buttons_frame = ttk.Frame(self.stock_tab)
        buttons_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(buttons_frame, text="Generate Report", command=self.generate_stock_report,
                  style='primary.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(buttons_frame, text="Export to PDF", command=self.export_stock_pdf,
                  style='danger.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(buttons_frame, text="Export to Excel", command=self.export_stock_excel,
                  style='success.TButton').pack(side=tk.LEFT, padx=5)
        
        # Results frame
        results_frame = ttk.Frame(self.stock_tab)
        results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Treeview
        self.stock_tree = ttk.Treeview(results_frame)
        self.stock_tree.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=self.stock_tree.yview)
        self.stock_tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Summary frame
        summary_frame = ttk.Frame(self.stock_tab)
        summary_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(summary_frame, text="Total Stock Value:").pack(side=tk.LEFT, padx=5)
        self.total_stock_var = tk.StringVar(value="$0.00")
        ttk.Label(summary_frame, textvariable=self.total_stock_var, font=('Helvetica', 12, 'bold')).pack(side=tk.LEFT, padx=5)
        
    def load_brands(self):
        try:
            conn = create_connection('data/phone_store.db')
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT brand FROM phones ORDER BY brand")
            brands = [row[0] for row in cursor.fetchall()]
            conn.close()
            
            self.stock_brand['values'] = brands
            if brands:
                self.stock_brand.current(0)
                
        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to load brands: {str(e)}")
    
    def on_stock_filter_change(self, event):
        if self.stock_filter.get() == 'By Brand':
            self.stock_brand['state'] = 'readonly'
        else:
            self.stock_brand['state'] = 'disabled'
    
    def generate_sales_report(self):
        date_from = self.sales_date_from.entry.get()
        date_to = self.sales_date_to.entry.get()
        group_by = self.sales_group_by.get()
        
        if not date_from or not date_to:
            messagebox.showwarning("Warning", "Please select both start and end dates")
            return
            
        try:
            conn = create_connection('data/phone_store.db')
            cursor = conn.cursor()
            
            # Determine group by SQL
            group_by_sql = ""
            select_fields = []
            
            if group_by == 'Day':
                group_by_sql = "strftime('%Y-%m-%d', s.sale_date)"
                select_fields = ["Date as day", "COUNT(*) as sales_count", "SUM(s.quantity) as total_quantity", 
                               "SUM(s.total_price) as total_sales"]
            elif group_by == 'Week':
                group_by_sql = "strftime('%Y-%W', s.sale_date)"
                select_fields = ["'Week ' || strftime('%W', s.sale_date) || ' ' || strftime('%Y', s.sale_date) as week", 
                                "COUNT(*) as sales_count", "SUM(s.quantity) as total_quantity", 
                                "SUM(s.total_price) as total_sales"]
            elif group_by == 'Month':
                group_by_sql = "strftime('%Y-%m', s.sale_date)"
                select_fields = ["strftime('%Y-%m', s.sale_date) as month", "COUNT(*) as sales_count", 
                               "SUM(s.quantity) as total_quantity", "SUM(s.total_price) as total_sales"]
            elif group_by == 'Year':
                group_by_sql = "strftime('%Y', s.sale_date)"
                select_fields = ["strftime('%Y', s.sale_date) as year", "COUNT(*) as sales_count", 
                               "SUM(s.quantity) as total_quantity", "SUM(s.total_price) as total_sales"]
            elif group_by == 'Product':
                group_by_sql = "p.id"
                select_fields = ["p.brand || ' ' || p.model as product", "COUNT(*) as sales_count", 
                               "SUM(s.quantity) as total_quantity", "SUM(s.total_price) as total_sales"]
            elif group_by == 'Payment Method':
                group_by_sql = "s.payment_method"
                select_fields = ["s.payment_method", "COUNT(*) as sales_count", 
                               "SUM(s.quantity) as total_quantity", "SUM(s.total_price) as total_sales"]
            
            # Build query
            query = f"""
            SELECT {', '.join(select_fields)}
            FROM sales s
            LEFT JOIN phones p ON s.phone_id = p.id
            WHERE s.sale_date BETWEEN ? AND ?
            GROUP BY {group_by_sql}
            ORDER BY {group_by_sql}
            """
            
            cursor.execute(query, (date_from, date_to))
            results = cursor.fetchall()
            
            # Get total sales
            cursor.execute("""
            SELECT SUM(total_price) 
            FROM sales 
            WHERE sale_date BETWEEN ? AND ?
            """, (date_from, date_to))
            total_sales = cursor.fetchone()[0] or 0
            self.total_sales_var.set(f"${total_sales:,.2f}")
            
            conn.close()
            
            # Clear treeview
            for item in self.sales_tree.get_children():
                self.sales_tree.delete(item)
                
            # Configure columns based on group by
            self.sales_tree['columns'] = [f'col{i}' for i in range(len(select_fields))]
            for i, field in enumerate(select_fields):
                col_name = field.split(' as ')[-1].replace('_', ' ').title()
                self.sales_tree.heading(f'col{i}', text=col_name)
                self.sales_tree.column(f'col{i}', width=150, anchor=tk.CENTER if i > 0 else tk.W)
            
            # Insert data
            for row in results:
                self.sales_tree.insert('', tk.END, values=row)
                
        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to generate sales report: {str(e)}")
    
    def generate_stock_report(self):
        filter_type = self.stock_filter.get()
        
        try:
            conn = create_connection('data/phone_store.db')
            cursor = conn.cursor()
            
            # Build query based on filter
            query = """
            SELECT p.id, p.brand, p.model, p.imei, p.price, p.quantity, (p.price * p.quantity) as total_value
            FROM phones p
            """
            params = []
            
            if filter_type == 'Low Stock (<5)':
                query += " WHERE p.quantity < 5"
            elif filter_type == 'By Brand':
                brand = self.stock_brand.get()
                if brand:
                    query += " WHERE p.brand = ?"
                    params.append(brand)
            
            query += " ORDER BY p.brand, p.model"
            
            cursor.execute(query, params)
            results = cursor.fetchall()
            
            # Get total stock value
            cursor.execute("""
            SELECT SUM(price * quantity) 
            FROM phones
            """)
            total_stock = cursor.fetchone()[0] or 0
            self.total_stock_var.set(f"${total_stock:,.2f}")
            
            conn.close()
            
            # Clear treeview
            for item in self.stock_tree.get_children():
                self.stock_tree.delete(item)
                
            # Configure columns
            columns = ['ID', 'Brand', 'Model', 'IMEI', 'Price', 'Quantity', 'Total Value']
            self.stock_tree['columns'] = [f'col{i}' for i in range(len(columns))]
            for i, col in enumerate(columns):
                self.stock_tree.heading(f'col{i}', text=col)
                self.stock_tree.column(f'col{i}', width=120, anchor=tk.CENTER if i > 3 else tk.W)
            
            # Insert data
            for row in results:
                formatted_row = list(row[:5]) + [row[5], f"${row[6]:,.2f}"]
                self.stock_tree.insert('', tk.END, values=formatted_row)
                
        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Failed to generate stock report: {str(e)}")
    
    def export_sales_pdf(self):
        if not self.sales_tree.get_children():
            messagebox.showwarning("Warning", "No data to export")
            return
            
        try:
            # Get data from treeview
            columns = [self.sales_tree.heading(col)['text'] for col in self.sales_tree['columns']]
            data = []
            for item in self.sales_tree.get_children():
                data.append(self.sales_tree.item(item)['values'])
            
            # Create PDF
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            
            # Title
            pdf.cell(200, 10, txt="Sales Report", ln=1, align='C')
            pdf.ln(10)
            
            # Table header
            col_widths = [40, 30, 30, 30]  # Adjust based on your columns
            for i, col in enumerate(columns):
                pdf.cell(col_widths[i], 10, txt=col, border=1, align='C')
            pdf.ln()
            
            # Table data
            for row in data:
                for i, item in enumerate(row):
                    if isinstance(item, float):
                        txt = f"${item:,.2f}"
                    else:
                        txt = str(item)
                    pdf.cell(col_widths[i], 10, txt=txt, border=1, align='C' if i > 0 else 'L')
                pdf.ln()
            
            # Total
            pdf.ln(5)
            pdf.cell(100, 10, txt=f"Total Sales: {self.total_sales_var.get()}", ln=1)
            
            # Save file
            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF Files", "*.pdf")],
                initialfile="sales_report.pdf"
            )
            
            if file_path:
                pdf.output(file_path)
                messagebox.showinfo("Success", f"Report saved to {file_path}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export PDF: {str(e)}")
    
    def export_sales_excel(self):
        if not self.sales_tree.get_children():
            messagebox.showwarning("Warning", "No data to export")
            return
            
        try:
            # Get data from treeview
            columns = [self.sales_tree.heading(col)['text'] for col in self.sales_tree['columns']]
            data = []
            for item in self.sales_tree.get_children():
                data.append(self.sales_tree.item(item)['values'])
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sales Report"
            
            # Write header
            for col_num, col_name in enumerate(columns, 1):
                ws.cell(row=1, column=col_num, value=col_name).font = Font(bold=True)
            
            # Write data
            for row_num, row_data in enumerate(data, 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=cell_value)
            
            # Auto-fit columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
            
            # Save file
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile="sales_report.xlsx"
            )
            
            if file_path:
                wb.save(file_path)
                messagebox.showinfo("Success", f"Report saved to {file_path}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export Excel: {str(e)}")
    
    def export_stock_pdf(self):
        if not self.stock_tree.get_children():
            messagebox.showwarning("Warning", "No data to export")
            return
            
        try:
            # Get data from treeview
            columns = [self.stock_tree.heading(col)['text'] for col in self.stock_tree['columns']]
            data = []
            for item in self.stock_tree.get_children():
                data.append(self.stock_tree.item(item)['values'])
            
            # Create PDF
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            
            # Title
            title = "Stock Report"
            if self.stock_filter.get() == 'Low Stock (<5)':
                title = "Low Stock Report"
            elif self.stock_filter.get() == 'By Brand':
                title = f"Stock Report - {self.stock_brand.get()}"
                
            pdf.cell(200, 10, txt=title, ln=1, align='C')
            pdf.ln(10)
            
            # Table header
            col_widths = [15, 30, 40, 50, 25, 25, 30]  # Adjust based on your columns
            for i, col in enumerate(columns):
                pdf.cell(col_widths[i], 10, txt=col, border=1, align='C')
            pdf.ln()
            
            # Table data
            for row in data:
                for i, item in enumerate(row):
                    pdf.cell(col_widths[i], 10, txt=str(item), border=1, align='C' if i > 2 else 'L')
                pdf.ln()
            
            # Total
            pdf.ln(5)
            pdf.cell(100, 10, txt=f"Total Stock Value: {self.total_stock_var.get()}", ln=1)
            
            # Save file
            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF Files", "*.pdf")],
                initialfile="stock_report.pdf"
            )
            
            if file_path:
                pdf.output(file_path)
                messagebox.showinfo("Success", f"Report saved to {file_path}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export PDF: {str(e)}")
    
    def export_stock_excel(self):
        if not self.stock_tree.get_children():
            messagebox.showwarning("Warning", "No data to export")
            return
            
        try:
            # Get data from treeview
            columns = [self.stock_tree.heading(col)['text'] for col in self.stock_tree['columns']]
            data = []
            for item in self.stock_tree.get_children():
                data.append(self.stock_tree.item(item)['values'])
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Stock Report"
            
            # Title
            title = "Stock Report"
            if self.stock_filter.get() == 'Low Stock (<5)':
                title = "Low Stock Report"
            elif self.stock_filter.get() == 'By Brand':
                title = f"Stock Report - {self.stock_brand.get()}"
            
            ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
            
            # Write header
            for col_num, col_name in enumerate(columns, 1):
                ws.cell(row=2, column=col_num, value=col_name).font = Font(bold=True)
            
            # Write data
            for row_num, row_data in enumerate(data, 3):
                for col_num, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=cell_value)
            
            # Add total
            ws.cell(row=len(data)+3, column=len(columns)-1, value="Total Value:").font = Font(bold=True)
            ws.cell(row=len(data)+3, column=len(columns), value=self.total_stock_var.get()).font = Font(bold=True)
            
            # Auto-fit columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
            
            # Save file
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile="stock_report.xlsx"
            )
            
            if file_path:
                wb.save(file_path)
                messagebox.showinfo("Success", f"Report saved to {file_path}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export Excel: {str(e)}")